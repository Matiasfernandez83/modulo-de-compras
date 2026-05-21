from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
import os
from datetime import datetime

def get_db():
    """Obtener conexión a la base de datos"""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_path = os.path.join(base_dir, 'database', 'gestion_compras.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def export_lista_precios_to_excel(lista_id, output_path):
    """
    Exporta una lista de precios a Excel con formato.
    
    Args:
        lista_id: ID de la lista de precios
        output_path: Ruta donde guardar el archivo
    
    Returns:
        str: Ruta del archivo generado
    """
    conn = get_db()
    cursor = conn.cursor()
    
    # Obtener datos de la lista
    cursor.execute("""
        SELECT lp.*, p.nombre as proveedor_nombre
        FROM listas_precios lp
        JOIN proveedores p ON lp.proveedor_id = p.id
        WHERE lp.id = ?
    """, (lista_id,))
    lista = cursor.fetchone()
    
    if not lista:
        conn.close()
        raise ValueError("Lista de precios no encontrada")
    
    # Obtener items
    cursor.execute("""
        SELECT lpi.*, a.codigo_interno, a.nombre as articulo_nombre, c.nombre as categoria_nombre
        FROM lista_precios_items lpi
        JOIN articulos a ON lpi.articulo_id = a.id
        LEFT JOIN categorias c ON a.categoria_id = c.id
        WHERE lpi.lista_precio_id = ?
        ORDER BY c.nombre, a.nombre
    """, (lista_id,))
    items = cursor.fetchall()
    conn.close()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezado de la lista
    ws['A1'] = "LISTA DE PRECIOS"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Proveedor: {lista['proveedor_nombre']}"
    ws['A3'] = f"Nombre: {lista['nombre']}"
    ws['A4'] = f"Moneda: {lista['moneda']}"
    ws['A5'] = f"Vigencia: {lista['fecha_vigencia']}"
    ws['A6'] = f"Fecha de Ingreso: {lista['fecha_ingreso']}"
    
    # Encabezados de columnas
    headers = [
        'Código Interno',
        'Código Proveedor',
        'Artículo',
        'Categoría',
        'Presentación',
        'Precio Bruto',
        'IVA %',
        'Descuento %',
        'Precio Neto'
    ]
    
    row = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    row = 9
    for item in items:
        ws.cell(row=row, column=1, value=item['codigo_interno']).border = border
        ws.cell(row=row, column=2, value=item['codigo_proveedor']).border = border
        ws.cell(row=row, column=3, value=item['articulo_nombre']).border = border
        ws.cell(row=row, column=4, value=item['categoria_nombre'] or '').border = border
        ws.cell(row=row, column=5, value=item['presentacion'] or '').border = border
        ws.cell(row=row, column=6, value=item['precio_bruto']).border = border
        ws.cell(row=row, column=6, value=item['precio_bruto']).number_format = '$#,##0.00'
        ws.cell(row=row, column=7, value=item['iva_porcentaje']).border = border
        ws.cell(row=row, column=8, value=item['descuento_porcentaje']).border = border
        ws.cell(row=row, column=9, value=item['precio_neto']).border = border
        ws.cell(row=row, column=9, value=item['precio_neto']).number_format = '$#,##0.00'
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Guardar
    wb.save(output_path)
    return output_path

def export_competencia_to_excel(competencia_id, output_path):
    """
    Exporta una competencia a Excel con formato de comparación.
    
    Args:
        competencia_id: ID de la competencia
        output_path: Ruta donde guardar el archivo
    
    Returns:
        str: Ruta del archivo generado
    """
    conn = get_db()
    cursor = conn.cursor()
    
    # Obtener competencia
    cursor.execute("SELECT * FROM competencias WHERE id = ?", (competencia_id,))
    competencia = cursor.fetchone()
    
    if not competencia:
        conn.close()
        raise ValueError("Competencia no encontrada")
    
    # Obtener proveedores participantes
    cursor.execute("""
        SELECT DISTINCT p.id, p.nombre
        FROM proveedores p
        JOIN competencia_items ci ON ci.proveedor_id = p.id
        WHERE ci.competencia_id = ?
    """, (competencia_id,))
    proveedores = cursor.fetchall()
    
    # Obtener items de la competencia
    cursor.execute("""
        SELECT ci.*, a.codigo_interno, a.nombre as articulo_nombre, p.nombre as proveedor_nombre
        FROM competencia_items ci
        JOIN articulos a ON ci.articulo_id = a.id
        JOIN proveedores p ON ci.proveedor_id = p.id
        WHERE ci.competencia_id = ?
        ORDER BY a.nombre, p.nombre
    """, (competencia_id,))
    items = cursor.fetchall()
    conn.close()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Competencia"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    
    # Título
    ws['A1'] = f"COMPETENCIA: {competencia['nombre']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Fecha: {competencia['fecha_creacion']}"
    ws['A3'] = f"Estado: {competencia['estado']}"
    
    # Encabezados
    row = 5
    ws.cell(row=row, column=1, value="Artículo").fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    
    for col, prov in enumerate(proveedores, 2):
        cell = ws.cell(row=row, column=col, value=prov['nombre'])
        cell.fill = header_fill
        cell.font = header_font
    
    # Agrupar items por artículo
    articulos_dict = {}
    for item in items:
        art_id = item['articulo_id']
        if art_id not in articulos_dict:
            articulos_dict[art_id] = {
                'nombre': item['articulo_nombre'],
                'precios': {}
            }
        articulos_dict[art_id]['precios'][item['proveedor_id']] = item['precio_neto']
    
    # Escribir datos
    row = 6
    for art_id, art_data in articulos_dict.items():
        ws.cell(row=row, column=1, value=art_data['nombre'])
        
        precios = list(art_data['precios'].values())
        min_precio = min(precios) if precios else 0
        max_precio = max(precios) if precios else 0
        
        for col, prov in enumerate(proveedores, 2):
            precio = art_data['precios'].get(prov['id'])
            if precio:
                cell = ws.cell(row=row, column=col, value=precio)
                cell.number_format = '$#,##0.00'
                
                # Aplicar color según precio
                if precio == min_precio:
                    cell.fill = green_fill
                elif precio == max_precio:
                    cell.fill = red_fill
                else:
                    cell.fill = yellow_fill
        
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 40
    for col in range(2, len(proveedores) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    wb.save(output_path)
    return output_path

def export_orden_compra_to_excel(orden_id, output_path):
    """
    Exporta una orden de compra a Excel.
    """
    conn = get_db()
    cursor = conn.cursor()
    
    # Obtener orden
    cursor.execute("""
        SELECT c.*, p.nombre as proveedor_nombre, p.direccion, p.telefono
        FROM comprobantes c
        JOIN proveedores p ON c.proveedor_id = p.id
        WHERE c.id = ? AND c.tipo = 'orden_compra'
    """, (orden_id,))
    orden = cursor.fetchone()
    
    if not orden:
        conn.close()
        raise ValueError("Orden de compra no encontrada")
    
    # Obtener items
    cursor.execute("""
        SELECT ci.*, a.codigo_interno, a.nombre as articulo_nombre
        FROM comprobante_items ci
        JOIN articulos a ON ci.articulo_id = a.id
        WHERE ci.comprobante_id = ?
    """, (orden_id,))
    items = cursor.fetchall()
    conn.close()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orden de Compra"
    
    # Título
    ws['A1'] = "ORDEN DE COMPRA"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    
    ws['A3'] = f"Número: {orden['numero']}"
    ws['A4'] = f"Fecha: {orden['fecha_emision']}"
    ws['A5'] = f"Proveedor: {orden['proveedor_nombre']}"
    ws['A6'] = f"Dirección: {orden['direccion']}"
    ws['A7'] = f"Teléfono: {orden['telefono']}"
    ws['A8'] = f"Estado: {orden['estado']}"
    
    # Encabezados de items
    headers = ['Código', 'Artículo', 'Cantidad', 'Precio Unit.', 'Subtotal']
    row = 10
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
    
    # Items
    row = 11
    total = 0
    for item in items:
        ws.cell(row=row, column=1, value=item['codigo_interno'])
        ws.cell(row=row, column=2, value=item['articulo_nombre'])
        ws.cell(row=row, column=3, value=item['cantidad'])
        ws.cell(row=row, column=4, value=item['precio_unitario']).number_format = '$#,##0.00'
        subtotal = item['cantidad'] * item['precio_unitario']
        ws.cell(row=row, column=5, value=subtotal).number_format = '$#,##0.00'
        total += subtotal
        row += 1
    
    # Total
    row += 1
    ws.cell(row=row, column=4, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=row, column=5, value=total).number_format = '$#,##0.00'
    ws.cell(row=row, column=5).font = Font(bold=True)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    wb.save(output_path)
    return output_path

def export_proveedores_to_excel(output_path):
    """Exporta todos los proveedores activos a Excel."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM proveedores WHERE activo = 1 ORDER BY nombre")
    proveedores = cursor.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Teléfono', 'Dirección', 'Email', 'Forma Pago', 'Plazo Pago', 'Tiempo Entrega']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header).font = Font(bold=True)
    
    # Datos
    for row, prov in enumerate(proveedores, 2):
        ws.cell(row, 1, prov['id'])
        ws.cell(row, 2, prov['nombre'])
        ws.cell(row, 3, prov['telefono'])
        ws.cell(row, 4, prov['direccion'])
        ws.cell(row, 5, prov['email'])
        ws.cell(row, 6, prov['forma_pago'])
        ws.cell(row, 7, prov['plazo_pago'])
        ws.cell(row, 8, prov['tiempo_entrega'])
    
    wb.save(output_path)
    return output_path

def export_articulos_to_excel(output_path):
    """Exporta todos los artículos activos a Excel."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT a.*, c.nombre as categoria_nombre
        FROM articulos a
        LEFT JOIN categorias c ON a.categoria_id = c.id
        WHERE a.activo = 1
        ORDER BY c.nombre, a.nombre
    """)
    articulos = cursor.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Artículos"
    
    # Encabezados
    headers = ['Código', 'Nombre', 'Descripción', 'Categoría', 'Unidad Medida']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header).font = Font(bold=True)
    
    # Datos
    for row, art in enumerate(articulos, 2):
        ws.cell(row, 1, art['codigo_interno'])
        ws.cell(row, 2, art['nombre'])
        ws.cell(row, 3, art['descripcion'])
        ws.cell(row, 4, art['categoria_nombre'] or '')
        ws.cell(row, 5, art['unidad_medida'])
    
    wb.save(output_path)
    return output_path

def export_usuarios_to_excel(output_path):
    """Exporta todos los usuarios activos a Excel."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, email, nombre_completo, rol, ultimo_acceso FROM usuarios WHERE activo = 1 ORDER BY username")
    usuarios = cursor.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Encabezados
    headers = ['ID', 'Usuario', 'Email', 'Nombre Completo', 'Rol', 'Último Acceso']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header).font = Font(bold=True)
    
    # Datos
    for row, user in enumerate(usuarios, 2):
        ws.cell(row, 1, user['id'])
        ws.cell(row, 2, user['username'])
        ws.cell(row, 3, user['email'])
        ws.cell(row, 4, user['nombre_completo'])
        ws.cell(row, 5, user['rol'])
        ws.cell(row, 6, user['ultimo_acceso'])
    
    wb.save(output_path)
    return output_path

