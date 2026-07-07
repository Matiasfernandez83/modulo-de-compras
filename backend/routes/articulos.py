from flask import Blueprint, request, jsonify
from middleware.session_auth import login_required, get_current_user_id, require_permission
import sqlite3
from routes.audit import registrar_auditoria

articulos_bp = Blueprint('articulos', __name__)

import os

from database.connection import get_db


def _derivar_prefijo(nombre):
    """Prefijo de 3 letras a partir del nombre (estilo Polaris).

    'FRUTAS Y VERDURAS' -> 'FYV' (iniciales), 'FIAMBRES' -> 'FIA' (primeras 3 letras).
    """
    palabras = [p for p in ''.join(c if c.isalpha() or c.isspace() else ' ' for c in nombre).split() if p]
    if len(palabras) >= 3:
        return ''.join(p[0] for p in palabras[:3]).upper()
    letras = ''.join(palabras)
    return letras[:3].upper().ljust(3, 'X')


def _generar_codigo(cursor, categoria_id, subcategoria_id):
    """Generar código estilo Polaris: PREFIJO_RUBRO + PREFIJO_SUBRUBRO + secuencial.

    Ej: FIA + FFF + 001 = FIAFFF001. Devuelve (codigo, error).
    """
    cat = cursor.execute("SELECT prefijo, nombre FROM categorias WHERE id = ?", (categoria_id,)).fetchone()
    if not cat:
        return None, 'Categoría inexistente'
    sub = cursor.execute(
        "SELECT prefijo FROM subcategorias WHERE id = ? AND categoria_id = ?",
        (subcategoria_id, categoria_id)
    ).fetchone()
    if not sub:
        return None, 'Subcategoría inexistente o no pertenece a la categoría'

    prefijo_cat = cat['prefijo'] or _derivar_prefijo(cat['nombre'])
    prefijo = (prefijo_cat + sub['prefijo']).upper()

    # Buscar el número más alto ya usado con este prefijo
    filas = cursor.execute(
        "SELECT codigo_interno FROM articulos WHERE codigo_interno LIKE ?", (prefijo + '%',)
    ).fetchall()
    maximo = 0
    for fila in filas:
        sufijo = fila['codigo_interno'][len(prefijo):]
        if sufijo.isdigit():
            maximo = max(maximo, int(sufijo))
    return f"{prefijo}{maximo + 1:03d}", None


@articulos_bp.route('', methods=['GET'])
@login_required
def get_articulos():
    """Obtener todos los artículos"""
    conn = get_db()
    cursor = conn.cursor()

    categoria_id = request.args.get('categoria_id')
    activo = request.args.get('activo', '1')
    # Paginación opcional: sin ?page= se devuelve la lista completa (compatible con el frontend actual)
    page = request.args.get('page', type=int)
    per_page = min(request.args.get('per_page', 50, type=int), 200)

    where = "WHERE a.activo = ?"
    params = [activo]
    if categoria_id:
        where = "WHERE a.categoria_id = ? AND a.activo = ?"
        params = [categoria_id, activo]

    query = f"""
        SELECT a.id, a.codigo_interno, a.codigo_softland, a.nombre, a.descripcion,
               a.unidad_medida, a.activo, a.categoria_id, a.subcategoria_id,
               c.nombre as categoria_nombre, s.nombre as subcategoria_nombre
        FROM articulos a
        LEFT JOIN categorias c ON a.categoria_id = c.id
        LEFT JOIN subcategorias s ON a.subcategoria_id = s.id
        {where}
        ORDER BY a.nombre
    """

    if page:
        cursor.execute(f"SELECT COUNT(*) FROM articulos a {where}", params)
        total = cursor.fetchone()[0]
        cursor.execute(query + " LIMIT ? OFFSET ?", (*params, per_page, (page - 1) * per_page))
        articulos = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'items': articulos, 'total': total, 'page': page, 'per_page': per_page}), 200

    cursor.execute(query, params)
    articulos = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return jsonify(articulos), 200

@articulos_bp.route('/<int:id>', methods=['GET'])
@login_required
def get_articulo(id):
    """Obtener un artículo por ID"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT a.id, a.codigo_interno, a.codigo_softland, a.nombre, a.descripcion,
               a.categoria_id, a.subcategoria_id, a.unidad_medida, a.activo,
               c.nombre as categoria_nombre, s.nombre as subcategoria_nombre
        FROM articulos a
        LEFT JOIN categorias c ON a.categoria_id = c.id
        LEFT JOIN subcategorias s ON a.subcategoria_id = s.id
        WHERE a.id = ?
    """, (id,))
    
    articulo = cursor.fetchone()
    conn.close()
    
    if not articulo:
        return jsonify({'error': 'Artículo no encontrado'}), 404
    
    return jsonify(dict(articulo)), 200

@articulos_bp.route('', methods=['POST'])
@require_permission('articulos', 'crear')
def create_articulo():
    """Crear un nuevo artículo.

    El código interno se puede omitir: si vienen categoria_id y subcategoria_id,
    se genera automáticamente con el formato RUBRO+SUBRUBRO+número (ej: FIAFFF001).
    """
    data = request.get_json() or {}

    if not data.get('nombre'):
        return jsonify({'error': 'El nombre es requerido'}), 400

    codigo_manual = data.get('codigo_interno')
    if not codigo_manual and not (data.get('categoria_id') and data.get('subcategoria_id')):
        return jsonify({'error': 'Ingresá un código interno, o elegí categoría y subcategoría para generarlo automáticamente'}), 400

    conn = get_db()
    cursor = conn.cursor()

    # Hasta 5 intentos por si dos usuarios generan el mismo código a la vez
    for _ in range(5):
        codigo = codigo_manual
        if not codigo:
            codigo, error = _generar_codigo(cursor, data.get('categoria_id'), data.get('subcategoria_id'))
            if error:
                conn.close()
                return jsonify({'error': error}), 400
        try:
            cursor.execute("""
                INSERT INTO articulos (codigo_interno, nombre, descripcion, categoria_id, subcategoria_id, unidad_medida)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                codigo,
                data.get('nombre'),
                data.get('descripcion'),
                data.get('categoria_id'),
                data.get('subcategoria_id'),
                data.get('unidad_medida')
            ))
            break
        except sqlite3.IntegrityError:
            if codigo_manual:
                conn.close()
                return jsonify({'error': 'El código interno ya existe'}), 400
            # código generado ya usado por otro usuario: recalcular
    else:
        conn.close()
        return jsonify({'error': 'No se pudo generar un código único, reintentá'}), 500

    articulo_id = cursor.lastrowid

    # Inicializar stock en 0
    cursor.execute("""
        INSERT INTO stock (articulo_id, cantidad) VALUES (?, 0)
    """, (articulo_id,))

    conn.commit()
    conn.close()

    # Registrar auditoría
    registrar_auditoria(
        usuario_id=get_current_user_id(),
        tabla='articulos',
        registro_id=articulo_id,
        accion='crear',
        datos_nuevos={
            'codigo_interno': codigo,
            'nombre': data.get('nombre')
        }
    )

    return jsonify({'id': articulo_id, 'codigo_interno': codigo, 'message': 'Artículo creado exitosamente'}), 201

@articulos_bp.route('/<int:id>', methods=['PUT'])
@require_permission('articulos', 'editar')
def update_articulo(id):
    """Actualizar un artículo"""
    data = request.get_json() or {}

    conn = get_db()
    cursor = conn.cursor()

    # Obtener datos actuales: los campos no enviados conservan su valor
    cursor.execute("""
        SELECT codigo_interno, nombre, descripcion, categoria_id, subcategoria_id, unidad_medida
        FROM articulos WHERE id = ?
    """, (id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Artículo no encontrado'}), 404
    actual = dict(row)
    datos_anteriores = {k: actual[k] for k in ('codigo_interno', 'nombre')}

    campos = ['codigo_interno', 'nombre', 'descripcion', 'categoria_id', 'subcategoria_id', 'unidad_medida']
    valores = [data.get(campo, actual[campo]) for campo in campos]

    cursor.execute("""
        UPDATE articulos
        SET codigo_interno = ?, nombre = ?, descripcion = ?, categoria_id = ?, subcategoria_id = ?, unidad_medida = ?
        WHERE id = ?
    """, (*valores, id))

    conn.commit()
    conn.close()
    
    # Registrar auditoría
    registrar_auditoria(
        usuario_id=get_current_user_id(),
        tabla='articulos',
        registro_id=id,
        accion='editar',
        datos_anteriores=datos_anteriores,
        datos_nuevos={
            'codigo_interno': data.get('codigo_interno'),
            'nombre': data.get('nombre')
        }
    )
    
    return jsonify({'message': 'Artículo actualizado exitosamente'}), 200

# Rutas de Categorías
@articulos_bp.route('/categorias', methods=['GET'])
@login_required
def get_categorias():
    """Obtener todas las categorías"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id, nombre, descripcion, prefijo, activo FROM categorias WHERE activo = 1 ORDER BY nombre")
    categorias = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return jsonify(categorias), 200

@articulos_bp.route('/categorias', methods=['POST'])
@login_required
def create_categoria():
    """Crear una nueva categoría (rubro). El prefijo se deriva del nombre si no se envía."""
    data = request.get_json() or {}

    if not data.get('nombre'):
        return jsonify({'error': 'El nombre es requerido'}), 400

    prefijo = (data.get('prefijo') or _derivar_prefijo(data['nombre'])).strip().upper()
    if len(prefijo) != 3 or not prefijo.isalpha():
        return jsonify({'error': 'El prefijo debe ser exactamente 3 letras (ej: FIA, GOL, FYV)'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO categorias (nombre, descripcion, prefijo) VALUES (?, ?, ?)
    """, (data.get('nombre'), data.get('descripcion'), prefijo))

    categoria_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'id': categoria_id, 'prefijo': prefijo, 'message': 'Categoría creada exitosamente'}), 201


# Rutas de Subcategorías
@articulos_bp.route('/subcategorias', methods=['GET'])
@login_required
def get_subcategorias():
    """Obtener subcategorías (subrubros), opcionalmente filtradas por categoría"""
    conn = get_db()
    cursor = conn.cursor()

    categoria_id = request.args.get('categoria_id')
    if categoria_id:
        cursor.execute("""
            SELECT id, categoria_id, nombre, prefijo, activo FROM subcategorias
            WHERE categoria_id = ? AND activo = 1 ORDER BY nombre
        """, (categoria_id,))
    else:
        cursor.execute("""
            SELECT id, categoria_id, nombre, prefijo, activo FROM subcategorias
            WHERE activo = 1 ORDER BY nombre
        """)
    subcategorias = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return jsonify(subcategorias), 200


@articulos_bp.route('/subcategorias', methods=['POST'])
@login_required
def create_subcategoria():
    """Crear una subcategoría (subrubro). El prefijo se deriva del nombre si no se envía."""
    data = request.get_json() or {}

    if not data.get('nombre') or not data.get('categoria_id'):
        return jsonify({'error': 'Nombre y categoría son requeridos'}), 400

    prefijo = (data.get('prefijo') or _derivar_prefijo(data['nombre'])).strip().upper()
    if len(prefijo) != 3 or not prefijo.isalpha():
        return jsonify({'error': 'El prefijo debe ser exactamente 3 letras (ej: FFF, ALF, FRU)'}), 400

    conn = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO subcategorias (categoria_id, nombre, prefijo) VALUES (?, ?, ?)
        """, (data.get('categoria_id'), data.get('nombre'), prefijo))
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Ya existe esa subcategoría en la categoría, o la categoría no existe'}), 400

    subcategoria_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'id': subcategoria_id, 'prefijo': prefijo, 'message': 'Subcategoría creada exitosamente'}), 201


@articulos_bp.route('/proximo-codigo', methods=['GET'])
@login_required
def get_proximo_codigo():
    """Vista previa del próximo código a asignar para una categoría + subcategoría"""
    categoria_id = request.args.get('categoria_id', type=int)
    subcategoria_id = request.args.get('subcategoria_id', type=int)
    if not categoria_id or not subcategoria_id:
        return jsonify({'error': 'categoria_id y subcategoria_id son requeridos'}), 400

    conn = get_db()
    cursor = conn.cursor()
    codigo, error = _generar_codigo(cursor, categoria_id, subcategoria_id)
    conn.close()

    if error:
        return jsonify({'error': error}), 400
    return jsonify({'codigo': codigo}), 200

@articulos_bp.route('/<int:id>/historial-compras', methods=['GET'])
@login_required
def get_historial_compras_articulo(id):
    """Historial de OC donde aparece este artículo"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT c.numero, c.fecha_emision, c.estado,
               p.nombre as proveedor,
               ci.cantidad, ci.precio_unitario, ci.subtotal,
               c.id as comprobante_id
        FROM comprobante_items ci
        JOIN comprobantes c ON ci.comprobante_id = c.id
        JOIN proveedores p ON c.proveedor_id = p.id
        WHERE ci.articulo_id = ? AND c.tipo = 'orden_compra'
        ORDER BY c.fecha_emision DESC
        LIMIT 50
    """, (id,))
    compras = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(compras), 200

@articulos_bp.route('/<int:id>/historial-precios', methods=['GET'])
@login_required
def get_historial_precios_articulo(id):
    """Historial de listas de precios que incluyen este artículo"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT lp.nombre as lista, lp.fecha_vigencia, lp.moneda,
               p.nombre as proveedor,
               lpi.precio_bruto, lpi.descuento_porcentaje, lpi.iva_porcentaje, lpi.precio_neto,
               lp.id as lista_id
        FROM lista_precios_items lpi
        JOIN listas_precios lp ON lpi.lista_precio_id = lp.id
        JOIN proveedores p ON lp.proveedor_id = p.id
        WHERE lpi.articulo_id = ?
        ORDER BY lp.fecha_vigencia DESC
        LIMIT 50
    """, (id,))
    precios = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(precios), 200

@articulos_bp.route('/template', methods=['GET'])
@login_required
def download_template_articulos():
    """Descargar template Excel para carga masiva de artículos"""
    from flask import send_file
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artículos"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center")

    headers = ["Código Interno *", "Nombre *", "Descripción", "Categoría", "Unidad de Medida"]
    col_widths = [20, 35, 40, 20, 20]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    ejemplos = [
        ["ART001", "Harina 000", "Harina de trigo tipo 000", "SECOS", "KG"],
        ["ART002", "Azúcar", "Azúcar blanca refinada", "SECOS", "KG"],
        ["ART014", "Aceite Girasol", "Aceite de girasol 1L", "SECOS", "LT"],
    ]
    for row_data in ejemplos:
        ws.append(row_data)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf, as_attachment=True,
        download_name="template_articulos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
