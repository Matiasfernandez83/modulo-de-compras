from flask import Blueprint, request, jsonify
from middleware.session_auth import login_required, get_current_user_id, require_permission
import sqlite3
from routes.audit import registrar_auditoria

proveedores_bp = Blueprint('proveedores', __name__)

import os

from database.connection import get_db

@proveedores_bp.route('', methods=['GET'])
@login_required
def get_proveedores():
    """Obtener todos los proveedores"""
    conn = get_db()
    cursor = conn.cursor()

    activo = request.args.get('activo', '1')
    # Paginación opcional: sin ?page= se devuelve la lista completa (compatible con el frontend actual)
    page = request.args.get('page', type=int)
    per_page = min(request.args.get('per_page', 50, type=int), 200)

    query = """
        SELECT id, nombre, telefono, direccion, email, forma_pago, plazo_pago,
               tiempo_entrega, activo, fecha_creacion
        FROM proveedores
        WHERE activo = ?
        ORDER BY nombre
    """

    if page:
        cursor.execute("SELECT COUNT(*) FROM proveedores WHERE activo = ?", (activo,))
        total = cursor.fetchone()[0]
        cursor.execute(query + " LIMIT ? OFFSET ?", (activo, per_page, (page - 1) * per_page))
        proveedores = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'items': proveedores, 'total': total, 'page': page, 'per_page': per_page}), 200

    cursor.execute(query, (activo,))
    proveedores = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return jsonify(proveedores), 200

@proveedores_bp.route('/<int:id>', methods=['GET'])
@login_required
def get_proveedor(id):
    """Obtener un proveedor por ID"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, nombre, telefono, direccion, email, forma_pago, plazo_pago,
               tiempo_entrega, activo, fecha_creacion
        FROM proveedores WHERE id = ?
    """, (id,))
    
    proveedor = cursor.fetchone()
    conn.close()
    
    if not proveedor:
        return jsonify({'error': 'Proveedor no encontrado'}), 404
    
    return jsonify(dict(proveedor)), 200

@proveedores_bp.route('', methods=['POST'])
@require_permission('proveedores', 'crear')
def create_proveedor():
    """Crear un nuevo proveedor"""
    data = request.get_json()
    
    if not data.get('nombre'):
        return jsonify({'error': 'El nombre es requerido'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        INSERT INTO proveedores (nombre, telefono, direccion, email, forma_pago, plazo_pago, tiempo_entrega)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get('nombre'),
        data.get('telefono'),
        data.get('direccion'),
        data.get('email'),
        data.get('forma_pago'),
        data.get('plazo_pago'),
        data.get('tiempo_entrega')
    ))
    
    proveedor_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # Registrar auditoría
    registrar_auditoria(
        usuario_id=get_current_user_id(),
        tabla='proveedores',
        registro_id=proveedor_id,
        accion='crear',
        datos_nuevos={
            'nombre': data.get('nombre'),
            'telefono': data.get('telefono'),
            'email': data.get('email')
        }
    )
    
    return jsonify({'id': proveedor_id, 'message': 'Proveedor creado exitosamente'}), 201

@proveedores_bp.route('/<int:id>', methods=['PUT'])
@require_permission('proveedores', 'editar')
def update_proveedor(id):
    """Actualizar un proveedor"""
    data = request.get_json() or {}

    conn = get_db()
    cursor = conn.cursor()

    # Obtener datos actuales: los campos no enviados conservan su valor
    cursor.execute("""
        SELECT nombre, telefono, direccion, email, forma_pago, plazo_pago, tiempo_entrega
        FROM proveedores WHERE id = ?
    """, (id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Proveedor no encontrado'}), 404
    actual = dict(row)
    datos_anteriores = {k: actual[k] for k in ('nombre', 'telefono', 'email')}

    campos = ['nombre', 'telefono', 'direccion', 'email', 'forma_pago', 'plazo_pago', 'tiempo_entrega']
    valores = [data.get(campo, actual[campo]) for campo in campos]

    cursor.execute("""
        UPDATE proveedores
        SET nombre = ?, telefono = ?, direccion = ?, email = ?,
            forma_pago = ?, plazo_pago = ?, tiempo_entrega = ?
        WHERE id = ?
    """, (*valores, id))

    conn.commit()
    conn.close()
    
    # Registrar auditoría
    registrar_auditoria(
        usuario_id=get_current_user_id(),
        tabla='proveedores',
        registro_id=id,
        accion='editar',
        datos_anteriores=datos_anteriores,
        datos_nuevos={
            'nombre': data.get('nombre'),
            'telefono': data.get('telefono'),
            'email': data.get('email')
        }
    )
    
    return jsonify({'message': 'Proveedor actualizado exitosamente'}), 200

@proveedores_bp.route('/<int:id>', methods=['DELETE'])
@require_permission('proveedores', 'eliminar')
def delete_proveedor(id):
    """Eliminar (soft delete) un proveedor"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Obtener datos antes de eliminar
    cursor.execute("SELECT nombre FROM proveedores WHERE id = ?", (id,))
    row = cursor.fetchone()
    datos_anteriores = dict(row) if row else {}
    
    cursor.execute("UPDATE proveedores SET activo = 0 WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    
    # Registrar auditoría
    registrar_auditoria(
        usuario_id=get_current_user_id(),
        tabla='proveedores',
        registro_id=id,
        accion='eliminar',
        datos_anteriores=datos_anteriores
    )
    
    return jsonify({'message': 'Proveedor eliminado exitosamente'}), 200

@proveedores_bp.route('/template', methods=['GET'])
@login_required
def download_template_proveedores():
    """Descargar template Excel para carga masiva de proveedores"""
    from flask import send_file
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center")

    headers = ["Nombre *", "Teléfono", "Dirección", "Email", "Forma de Pago", "Plazo Pago (días)", "Tiempo Entrega (días)"]
    col_widths = [35, 18, 40, 35, 20, 20, 22]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    ejemplos = [
        ["Distribuidora Ejemplo S.A.", "011-4567-8901", "Av. Corrientes 1234", "ventas@ejemplo.com", "Cuenta Corriente", 30, 5],
        ["Proveedor Local SRL", "011-9876-5432", "Av. San Martín 456", "info@local.com", "Contado", 0, 2],
    ]
    for row_data in ejemplos:
        ws.append(row_data)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf, as_attachment=True,
        download_name="template_proveedores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
