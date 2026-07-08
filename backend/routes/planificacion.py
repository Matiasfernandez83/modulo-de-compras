from flask import Blueprint, request, jsonify, send_file
from middleware.session_auth import login_required, get_current_user_id, require_permission
import sqlite3
import os
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

planificacion_bp = Blueprint('planificacion', __name__)

from database.connection import get_db

@planificacion_bp.route('', methods=['GET'])
@login_required
def get_planificaciones():
    """Obtener todas las planificaciones"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT pl.*, pr.nombre as producto_nombre, pr.codigo as producto_codigo
        FROM planificaciones pl
        LEFT JOIN productos pr ON pl.producto_id = pr.id
        ORDER BY pl.fecha_creacion DESC
    """)
    planificaciones = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(planificaciones), 200

@planificacion_bp.route('/<int:plan_id>', methods=['GET'])
@login_required
def get_planificacion_detail(plan_id):
    """Obtener detalle de una planificación con sus ítems"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM planificaciones WHERE id = ?", (plan_id,))
    plan = cursor.fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Planificación no encontrada'}), 404

    plan_dict = dict(plan)

    # Obtener ítems con info del artículo
    cursor.execute("""
        SELECT pi.id, pi.articulo_id, pi.cantidad_requerida, pi.fecha_necesidad,
               a.codigo_interno, a.nombre, a.unidad_medida,
               c.nombre as categoria
        FROM planificacion_items pi
        JOIN articulos a ON pi.articulo_id = a.id
        LEFT JOIN categorias c ON a.categoria_id = c.id
        WHERE pi.planificacion_id = ?
        ORDER BY c.nombre, a.nombre
    """, (plan_id,))
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()

    plan_dict['items'] = items
    plan_dict['total_items'] = len(items)
    return jsonify(plan_dict), 200

@planificacion_bp.route('', methods=['POST'])
@login_required
def create_planificacion():
    """Crear nueva planificación.

    Si viene producto_id + cantidad_unidades, se hace la explosión de materiales
    (estilo MRP): cada insumo de la estructura del producto se agrega como ítem
    con cantidad = cantidad_por_unidad × unidades a fabricar.
    """
    data = request.get_json() or {}
    if not data.get('nombre'):
        return jsonify({'error': 'El nombre es requerido'}), 400

    producto_id = data.get('producto_id')
    cantidad_unidades = float(data.get('cantidad_unidades') or 1)
    if producto_id and cantidad_unidades <= 0:
        return jsonify({'error': 'La cantidad de unidades debe ser mayor a 0'}), 400

    conn = get_db()
    cursor = conn.cursor()

    if producto_id:
        producto = cursor.execute("SELECT id, nombre FROM productos WHERE id = ? AND activo = 1",
                                  (producto_id,)).fetchone()
        if not producto:
            conn.close()
            return jsonify({'error': 'Producto no encontrado'}), 404

    cursor.execute("""
        INSERT INTO planificaciones (nombre, descripcion, fecha_inicio, fecha_fin,
                                     producto_id, cantidad_unidades, usuario_creacion_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (data['nombre'], data.get('descripcion'), data.get('fecha_inicio'),
          data.get('fecha_fin'), producto_id, cantidad_unidades, get_current_user_id()))
    plan_id = cursor.lastrowid

    items_generados = 0
    if producto_id:
        # Explosión de materiales: receta × unidades a fabricar
        cursor.execute("""
            INSERT INTO planificacion_items (planificacion_id, articulo_id, cantidad_requerida, fecha_necesidad)
            SELECT ?, articulo_id, cantidad_por_unidad * ?, ?
            FROM producto_materiales WHERE producto_id = ?
        """, (plan_id, cantidad_unidades, data.get('fecha_inicio'), producto_id))
        items_generados = cursor.rowcount

    conn.commit()
    conn.close()
    return jsonify({'id': plan_id, 'items_generados': items_generados,
                    'message': 'Planificación creada'}), 201


def _calcular_necesidades(cursor, plan_id):
    """Necesidades de compra: requerido - stock disponible, con semáforo."""
    cursor.execute("""
        SELECT pi.articulo_id, SUM(pi.cantidad_requerida) as cantidad_requerida,
               a.codigo_interno, a.codigo_softland, a.nombre, a.unidad_medida,
               s.nombre as categoria,
               COALESCE(st.cantidad, 0) as stock_disponible
        FROM planificacion_items pi
        JOIN articulos a ON pi.articulo_id = a.id
        LEFT JOIN subcategorias s ON a.subcategoria_id = s.id
        LEFT JOIN stock st ON st.articulo_id = a.id
        WHERE pi.planificacion_id = ?
        GROUP BY pi.articulo_id
        ORDER BY s.nombre, a.nombre
    """, (plan_id,))
    necesidades = []
    for row in cursor.fetchall():
        item = dict(row)
        neta = max(0, (item['cantidad_requerida'] or 0) - (item['stock_disponible'] or 0))
        item['necesidad_neta'] = round(neta, 2)
        if neta <= 0:
            item['semaforo'] = 'verde'      # el stock cubre todo
        elif item['stock_disponible'] > 0:
            item['semaforo'] = 'amarillo'   # cubre una parte
        else:
            item['semaforo'] = 'rojo'       # hay que comprar todo
        necesidades.append(item)
    return necesidades


@planificacion_bp.route('/<int:plan_id>/necesidades', methods=['GET'])
@login_required
def get_necesidades(plan_id):
    """Cálculo de necesidades de compra de la planificación (requerido vs stock)"""
    conn = get_db()
    cursor = conn.cursor()

    plan = cursor.execute("""
        SELECT pl.*, pr.nombre as producto_nombre FROM planificaciones pl
        LEFT JOIN productos pr ON pl.producto_id = pr.id WHERE pl.id = ?
    """, (plan_id,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Planificación no encontrada'}), 404

    necesidades = _calcular_necesidades(cursor, plan_id)
    conn.close()

    total = len(necesidades)
    a_comprar = sum(1 for n in necesidades if n['necesidad_neta'] > 0)
    return jsonify({
        'planificacion': dict(plan),
        'necesidades': necesidades,
        'resumen': {'total_articulos': total, 'a_comprar': a_comprar,
                    'cubiertos_por_stock': total - a_comprar}
    }), 200


@planificacion_bp.route('/<int:plan_id>', methods=['DELETE'])
@login_required
def delete_planificacion(plan_id):
    """Eliminar una planificación con sus ítems.

    Las comparativas generadas desde ella se conservan (se les quita el vínculo).
    """
    conn = get_db()
    cursor = conn.cursor()
    if not cursor.execute("SELECT 1 FROM planificaciones WHERE id = ?", (plan_id,)).fetchone():
        conn.close()
        return jsonify({'error': 'Planificación no encontrada'}), 404

    cursor.execute("UPDATE competencias SET planificacion_id = NULL WHERE planificacion_id = ?", (plan_id,))
    cursor.execute("DELETE FROM planificacion_items WHERE planificacion_id = ?", (plan_id,))
    cursor.execute("DELETE FROM planificaciones WHERE id = ?", (plan_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Planificación eliminada'}), 200


@planificacion_bp.route('/<int:plan_id>/categorias', methods=['GET'])
@login_required
def get_categorias_plan(plan_id):
    """Categorías/rubros presentes en los ítems de la planificación.

    Para cada categoría informa cuántos artículos hay y cuántos proveedores
    la proveen (según listas de precios y mapeo de códigos), para que el
    usuario elija por qué categorías armar la comparativa.
    """
    conn = get_db()
    cursor = conn.cursor()

    if not cursor.execute("SELECT 1 FROM planificaciones WHERE id = ?", (plan_id,)).fetchone():
        conn.close()
        return jsonify({'error': 'Planificación no encontrada'}), 404

    cursor.execute("""
        SELECT s.id as subcategoria_id, s.nombre as categoria,
               COUNT(DISTINCT pi.articulo_id) as articulos,
               COUNT(DISTINCT prov.proveedor_id) as proveedores
        FROM planificacion_items pi
        JOIN articulos a ON pi.articulo_id = a.id
        LEFT JOIN subcategorias s ON a.subcategoria_id = s.id
        LEFT JOIN (
            SELECT lpi.articulo_id, lp.proveedor_id
            FROM lista_precios_items lpi
            JOIN listas_precios lp ON lpi.lista_precio_id = lp.id AND lp.activo = 1
            UNION
            SELECT articulo_id, proveedor_id FROM mapeo_codigos_proveedor
        ) prov ON prov.articulo_id = pi.articulo_id
        WHERE pi.planificacion_id = ?
        GROUP BY s.id, s.nombre
        ORDER BY s.nombre
    """, (plan_id,))
    categorias = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(categorias), 200


@planificacion_bp.route('/<int:plan_id>/generar-competencia', methods=['POST'])
@require_permission('competencia', 'crear')
def generar_competencia(plan_id):
    """Generar una comparativa (competencia) desde la planificación.

    Toma los ítems de la planificación de las categorías elegidas, crea la
    competencia vinculada a la planificación, carga sus ítems con la cantidad
    necesaria y auto-vincula los proveedores que proveen esos insumos.
    """
    data = request.get_json() or {}
    subcategoria_ids = data.get('subcategoria_ids') or []
    if not subcategoria_ids:
        return jsonify({'error': 'Elegí al menos una categoría'}), 400

    conn = get_db()
    cursor = conn.cursor()

    plan = cursor.execute("SELECT * FROM planificaciones WHERE id = ?", (plan_id,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Planificación no encontrada'}), 404

    placeholders = ','.join('?' * len(subcategoria_ids))
    # Ítems de la planificación en esas categorías, agrupados por artículo
    cursor.execute(f"""
        SELECT pi.articulo_id, SUM(pi.cantidad_requerida) as cantidad,
               COALESCE(st.cantidad, 0) as stock
        FROM planificacion_items pi
        JOIN articulos a ON pi.articulo_id = a.id
        LEFT JOIN stock st ON st.articulo_id = a.id
        WHERE pi.planificacion_id = ? AND a.subcategoria_id IN ({placeholders})
        GROUP BY pi.articulo_id
    """, (plan_id, *subcategoria_ids))
    items = cursor.fetchall()
    if not items:
        conn.close()
        return jsonify({'error': 'No hay artículos en las categorías elegidas'}), 400

    articulo_ids = [it['articulo_id'] for it in items]
    art_ph = ','.join('?' * len(articulo_ids))

    # Proveedores que proveen alguno de esos artículos (listas de precios o mapeo)
    cursor.execute(f"""
        SELECT DISTINCT proveedor_id FROM (
            SELECT lp.proveedor_id
            FROM lista_precios_items lpi
            JOIN listas_precios lp ON lpi.lista_precio_id = lp.id AND lp.activo = 1
            WHERE lpi.articulo_id IN ({art_ph})
            UNION
            SELECT proveedor_id FROM mapeo_codigos_proveedor WHERE articulo_id IN ({art_ph})
        )
    """, (*articulo_ids, *articulo_ids))
    proveedor_ids = [row['proveedor_id'] for row in cursor.fetchall()]

    nombre = data.get('nombre') or f"Comparativa - {plan['nombre']}"
    cursor.execute("""
        INSERT INTO competencias (nombre, planificacion_id, origen, estado, usuario_creacion_id)
        VALUES (?, ?, 'planificacion', 'borrador', ?)
    """, (nombre, plan_id, get_current_user_id()))
    competencia_id = cursor.lastrowid

    for it in items:
        neta = max(0, (it['cantidad'] or 0) - (it['stock'] or 0))
        cursor.execute("""
            INSERT INTO competencia_items
                (competencia_id, articulo_id, cantidad_necesaria, necesidad, stock_disponible, compra_sugerida)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (competencia_id, it['articulo_id'], it['cantidad'], it['cantidad'], it['stock'], neta))

    for prov_id in proveedor_ids:
        cursor.execute("""
            INSERT INTO competencia_proveedores (competencia_id, proveedor_id) VALUES (?, ?)
        """, (competencia_id, prov_id))

    conn.commit()
    conn.close()

    return jsonify({
        'competencia_id': competencia_id,
        'items': len(items),
        'proveedores': len(proveedor_ids),
        'message': 'Comparativa generada desde la planificación'
    }), 201


@planificacion_bp.route('/<int:plan_id>/necesidades/export', methods=['GET'])
@login_required
def export_necesidades(plan_id):
    """Exportar las necesidades de compra a Excel"""
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    conn = get_db()
    cursor = conn.cursor()
    plan = cursor.execute("SELECT * FROM planificaciones WHERE id = ?", (plan_id,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Planificación no encontrada'}), 404
    necesidades = _calcular_necesidades(cursor, plan_id)
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Necesidades de compra"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center")

    ws['A1'] = f"Necesidades de compra — {plan['nombre']}"
    ws['A1'].font = Font(bold=True, size=13)

    headers = ["Código", "Cód. Softland", "Artículo", "Categoría", "Unidad",
               "Requerido", "Stock", "A Comprar", "Estado"]
    widths = [14, 14, 55, 22, 10, 12, 12, 12, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    colores = {'rojo': 'C62828', 'amarillo': 'F9A825', 'verde': '2E7D32'}
    r = 4
    for n in necesidades:
        ws.cell(row=r, column=1, value=n['codigo_interno'])
        ws.cell(row=r, column=2, value=n['codigo_softland'])
        ws.cell(row=r, column=3, value=n['nombre'])
        ws.cell(row=r, column=4, value=n['categoria'])
        ws.cell(row=r, column=5, value=n['unidad_medida'])
        ws.cell(row=r, column=6, value=n['cantidad_requerida'])
        ws.cell(row=r, column=7, value=n['stock_disponible'])
        ws.cell(row=r, column=8, value=n['necesidad_neta'])
        estado = ws.cell(row=r, column=9, value=n['semaforo'].upper())
        estado.font = Font(bold=True, color='FFFFFF')
        estado.fill = PatternFill('solid', fgColor=colores[n['semaforo']])
        estado.alignment = center
        r += 1
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{r-1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"necesidades_plan_{plan_id}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@planificacion_bp.route('/<int:plan_id>/items', methods=['POST'])
@login_required
def add_item_planificacion(plan_id):
    """Agregar item a planificación"""
    data = request.get_json()
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO planificacion_items (planificacion_id, articulo_id, cantidad_requerida)
        VALUES (?, ?, ?)
    """, (plan_id, data['articulo_id'], data.get('cantidad_estimada', data.get('cantidad_requerida', 0))))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Item agregado'}), 201

@planificacion_bp.route('/<int:plan_id>/items/bulk', methods=['POST'])
@login_required
def bulk_import_items(plan_id):
    """Importar ítems de planificación desde Excel"""
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Solo se aceptan archivos Excel (.xlsx, .xls)'}), 400

    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        conn = get_db()
        cursor = conn.cursor()

        insertados = 0
        errores = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # saltar filas vacías
                continue

            codigo = str(row[0]).strip()
            cantidad = row[2] if len(row) > 2 and row[2] else 1

            # Buscar artículo por código
            cursor.execute("SELECT id, nombre FROM articulos WHERE codigo_interno = ? AND activo = 1", (codigo,))
            art = cursor.fetchone()
            if not art:
                errores.append(f"Código '{codigo}' no encontrado")
                continue

            cursor.execute("""
                INSERT INTO planificacion_items (planificacion_id, articulo_id, cantidad_requerida)
                VALUES (?, ?, ?)
            """, (plan_id, art['id'], cantidad))
            insertados += 1

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'insertados': insertados,
            'errores': errores
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@planificacion_bp.route('/template', methods=['GET'])
@login_required
def download_template():
    """Descargar template Excel para carga masiva de planificación"""
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planificación"

    # Estilo encabezado
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center")

    headers = ["Código Artículo *", "Nombre Artículo (referencia)", "Cantidad Requerida *", "Fecha Necesidad (YYYY-MM-DD)"]
    col_widths = [22, 35, 22, 30]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    # Filas de ejemplo
    ejemplos = [
        ["ART001", "Harina 000", 500, "2026-04-01"],
        ["ART002", "Azúcar", 200, "2026-04-01"],
        ["ART003", "Arroz", 300, "2026-04-15"],
    ]
    for row_data in ejemplos:
        ws.append(row_data)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="template_planificacion.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
