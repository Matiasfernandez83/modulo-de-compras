"""Productos y sus estructuras (BOM / lista de materiales).

Un producto es lo que la empresa fabrica (ej: una batea). Su estructura define
qué insumos y en qué cantidad lleva UNA unidad. La planificación usa esta
receta para calcular necesidades de compra (explosión de materiales, como el
MRP de SAP).
"""

from flask import Blueprint, request, jsonify, send_file
from middleware.session_auth import login_required, require_permission
import sqlite3
import io

from database.connection import get_db, IntegrityError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

productos_bp = Blueprint('productos', __name__)


@productos_bp.route('', methods=['GET'])
@login_required
def get_productos():
    """Listar productos con el tamaño de su estructura"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT p.id, p.codigo, p.nombre, p.descripcion, p.activo, p.fecha_creacion,
               COUNT(pm.id) as total_materiales
        FROM productos p
        LEFT JOIN producto_materiales pm ON pm.producto_id = p.id
        WHERE p.activo = 1
        GROUP BY p.id
        ORDER BY p.nombre
    """)
    productos = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(productos), 200


@productos_bp.route('', methods=['POST'])
@login_required
def create_producto():
    """Crear un producto"""
    data = request.get_json() or {}
    if not data.get('codigo') or not data.get('nombre'):
        return jsonify({'error': 'Código y nombre son requeridos'}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO productos (codigo, nombre, descripcion) VALUES (?, ?, ?)",
                       (data['codigo'].strip().upper(), data['nombre'], data.get('descripcion')))
    except IntegrityError:
        conn.close()
        return jsonify({'error': 'Ya existe un producto con ese código'}), 400
    producto_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': producto_id, 'message': 'Producto creado exitosamente'}), 201


@productos_bp.route('/<int:id>', methods=['GET'])
@login_required
def get_producto(id):
    """Obtener un producto con su estructura completa (BOM)"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM productos WHERE id = ?", (id,))
    producto = cursor.fetchone()
    if not producto:
        conn.close()
        return jsonify({'error': 'Producto no encontrado'}), 404

    cursor.execute("""
        SELECT pm.id, pm.articulo_id, pm.cantidad_por_unidad, pm.seccion,
               pm.observaciones, pm.plano,
               a.codigo_interno, a.codigo_softland, a.nombre as articulo_nombre,
               a.unidad_medida, s.nombre as categoria
        FROM producto_materiales pm
        JOIN articulos a ON pm.articulo_id = a.id
        LEFT JOIN subcategorias s ON a.subcategoria_id = s.id
        WHERE pm.producto_id = ?
        ORDER BY pm.seccion, a.nombre
    """, (id,))
    materiales = [dict(row) for row in cursor.fetchall()]
    conn.close()

    resultado = dict(producto)
    resultado['materiales'] = materiales
    resultado['total_materiales'] = len(materiales)
    return jsonify(resultado), 200


@productos_bp.route('/<int:id>/materiales', methods=['POST'])
@login_required
def add_material(id):
    """Agregar un insumo a la estructura del producto"""
    data = request.get_json() or {}
    if not data.get('articulo_id') or not data.get('cantidad_por_unidad'):
        return jsonify({'error': 'Artículo y cantidad por unidad son requeridos'}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO producto_materiales (producto_id, articulo_id, cantidad_por_unidad, seccion, observaciones)
            VALUES (?, ?, ?, ?, ?)
        """, (id, data['articulo_id'], data['cantidad_por_unidad'],
              data.get('seccion'), data.get('observaciones')))
    except IntegrityError:
        conn.close()
        return jsonify({'error': 'Ese artículo ya está en la estructura (editalo en lugar de agregarlo), o el producto no existe'}), 400
    conn.commit()
    conn.close()
    return jsonify({'message': 'Insumo agregado a la estructura'}), 201


@productos_bp.route('/<int:id>/materiales/<int:material_id>', methods=['PUT'])
@login_required
def update_material(id, material_id):
    """Modificar la cantidad por unidad de un insumo de la estructura"""
    data = request.get_json() or {}
    if not data.get('cantidad_por_unidad'):
        return jsonify({'error': 'cantidad_por_unidad es requerida'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE producto_materiales SET cantidad_por_unidad = ?
        WHERE id = ? AND producto_id = ?
    """, (data['cantidad_por_unidad'], material_id, id))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Insumo no encontrado en la estructura'}), 404
    conn.commit()
    conn.close()
    return jsonify({'message': 'Cantidad actualizada'}), 200


@productos_bp.route('/<int:id>/materiales/<int:material_id>', methods=['DELETE'])
@login_required
def delete_material(id, material_id):
    """Quitar un insumo de la estructura del producto"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM producto_materiales WHERE id = ? AND producto_id = ?",
                   (material_id, id))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Insumo no encontrado en la estructura'}), 404
    conn.commit()
    conn.close()
    return jsonify({'message': 'Insumo quitado de la estructura'}), 200


@productos_bp.route('/<int:id>/materiales/bulk', methods=['POST'])
@login_required
def import_materiales(id):
    """Importar la receta (BOM) de un producto desde Excel.

    Columnas esperadas: Código | Cantidad | Sección (opcional) | Observaciones (opcional).
    El código matchea por código interno (TAL...) o por código Softland.
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado'}), 500
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Solo se aceptan archivos Excel (.xlsx, .xls)'}), 400

    conn = get_db()
    cursor = conn.cursor()
    if not cursor.execute("SELECT 1 FROM productos WHERE id = ?", (id,)).fetchone():
        conn.close()
        return jsonify({'error': 'Producto no encontrado'}), 404

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        conn.close()
        return jsonify({'error': f'No se pudo leer el Excel: {e}'}), 400
    ws = wb.active

    insertados, actualizados, errores = 0, 0, []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        codigo = str(row[0]).strip()
        cantidad = row[1] if len(row) > 1 else None
        seccion = str(row[2]).strip() if len(row) > 2 and row[2] else None
        obs = str(row[3]).strip() if len(row) > 3 and row[3] else None
        try:
            cantidad = float(str(cantidad).replace(',', '.'))
        except (TypeError, ValueError):
            errores.append(f'Fila {i}: cantidad inválida')
            continue
        if cantidad <= 0:
            continue

        art = cursor.execute(
            "SELECT id FROM articulos WHERE codigo_interno = ? OR codigo_softland = ? LIMIT 1",
            (codigo, codigo)
        ).fetchone()
        if not art:
            errores.append(f"Fila {i}: código '{codigo}' no encontrado en el catálogo")
            continue

        existe = cursor.execute(
            "SELECT id FROM producto_materiales WHERE producto_id = ? AND articulo_id = ?",
            (id, art['id'])
        ).fetchone()
        if existe:
            cursor.execute("""
                UPDATE producto_materiales SET cantidad_por_unidad = ?, seccion = ?, observaciones = ?
                WHERE id = ?
            """, (cantidad, seccion, obs, existe['id']))
            actualizados += 1
        else:
            cursor.execute("""
                INSERT INTO producto_materiales (producto_id, articulo_id, cantidad_por_unidad, seccion, observaciones)
                VALUES (?, ?, ?, ?, ?)
            """, (id, art['id'], cantidad, seccion, obs))
            insertados += 1

    conn.commit()
    conn.close()
    return jsonify({
        'insertados': insertados, 'actualizados': actualizados,
        'errores': errores[:50], 'total_errores': len(errores)
    }), 200


@productos_bp.route('/template', methods=['GET'])
@login_required
def download_template_producto():
    """Template Excel para cargar la receta (BOM) de un producto"""
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receta"
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    headers = ["Código (interno o Softland) *", "Cantidad por unidad *", "Sección", "Observaciones"]
    widths = [30, 20, 25, 40]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = fill
        c.font = font
        c.alignment = center
        ws.column_dimensions[c.column_letter].width = w
    ws.append(["TALHIE001", 12, "CAÑOS Y TUBOS", "Tubo estructural"])
    ws.append(["370", 28, "CAÑOS Y TUBOS", "Caño 60"])
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="template_receta.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
